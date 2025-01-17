import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

def fill_header_color(sheet, header_row_index, new_columns):
    """填充表头颜色"""
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    try:
        header_row = sheet[header_row_index]
        for new_col, _ in new_columns:
            for cell in header_row:
                if str(cell.value).strip() == new_col:
                    cell.fill = yellow_fill
                    print(f"成功填充表头颜色：{new_col}")
                    break
            else:
                print(f"警告：在表头中找不到列 '{new_col}'，无法填充颜色。")
    except Exception as e:
        print(f"填充颜色失败：{e}")
        return False
    return True


def insert_columns(wb_path, original_sheet_name, insert_after_sheet_name, df):
    """使用 pandas 插入新列并保存到新工作表,直接操作df"""
    print("-" * 30)
    print(f"开始执行 insert_columns 函数")
    print(f"目标文件路径：{wb_path}, 原始工作表：{original_sheet_name}, 新工作表：{insert_after_sheet_name}")

    try:
        new_columns = [
            ("ISCM终端MAC地址-注册状态", "ISCM终端MAC地址"),
            ("精简型号", "设备名称"),
            ("目前在用型号2", "设备名称"),
            ("是否出库在用一致", "设备名称"),
            ("LOID（SN码）", "业务号码")
        ]

        for new_col, original_col in new_columns:
            if new_col not in df.columns:
                try:
                    insert_index = df.columns.get_loc(original_col) + 1
                    df.insert(insert_index, new_col, "")
                    print(f"成功插入列：{new_col}")
                except KeyError:
                    print(f"错误：找不到原始列 '{original_col}'，请检查 Excel 文件是否存在此列。")
                    return False
                except Exception as e:
                    print(f"插入列时发生未知错误：{e}")
                    return False

        print(f"insert_columns 函数执行完毕")
        print("-" * 30)
        return True

    except Exception as e:
        print(f"insert_columns 函数执行过程中发生未知错误：{e}")
        return False
    finally:
        print(f"insert_columns 函数执行完毕")
        print("-" * 30)
'''
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill


def fill_header_color(sheet, header_row_index, new_columns):
    """填充表头颜色"""
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    try:
        header_row = sheet[header_row_index]
        for new_col, _ in new_columns:
            for cell in header_row:
                # 关键修改：将单元格值强制转换为 str 类型
                if str(cell.value).strip() == new_col:  # 使用 str() 转换
                    cell.fill = yellow_fill
                    print(f"成功填充表头颜色：{new_col}")
                    break
            else:
                print(f"警告：在表头中找不到列 '{new_col}'，无法填充颜色。")
    except Exception as e:
        print(f"填充颜色失败：{e}")
        return False
    return True


def insert_columns(wb_path, sheet_name):
    """插入新列并填充表头颜色"""
    print("-" * 30)
    print(f"开始执行 insert_columns 函数")
    print(f"目标文件路径：{wb_path}, 目标工作表：{sheet_name}")

    try:
        new_columns = [
            ("ISCM终端MAC地址-注册状态", "ISCM终端MAC地址"),
            #("精简型号", "设备名称"),
            ("是否出库在用一致", "设备名称"),
            ("目前在用型号2", "设备名称"),
            #("是否出库在用一致", "设备名称"),
            ("精简型号", "设备名称"),
            ("LOID（SN码）", "业务号码")
        ]

        try:
            df = pd.read_excel(wb_path, sheet_name=sheet_name, engine='openpyxl')
        except Exception as e:
            print(f"读取excel失败：{e}")
            return False

        for new_col, original_col in new_columns:
            if new_col not in df.columns:
                try:
                    insert_index = df.columns.get_loc(original_col) + 1
                    df.insert(insert_index, new_col, "")
                    print(f"成功插入列：{new_col}")
                except KeyError as e:
                    print(f"找不到原始列{original_col},请检查excel文件是否存在此列")
                    return False
                except Exception as e:
                    print(f"插入列失败{e}")
                    return False

        try:
            with pd.ExcelWriter(wb_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            print(f"保存excel失败:{e}")
            return False

        try:
            wb = openpyxl.load_workbook(wb_path)
            sheet = wb[sheet_name]

            header_row = sheet[1]
            if not fill_header_color(sheet, 1, new_columns):
                print("填充表头颜色失败。")
        except KeyError as e:
            print(f"未找到sheet:{sheet_name},请检查")
            return False
        except Exception as e:
            print(f"填充颜色失败：{e}")
            return False

        try:
            wb.save(wb_path)
        except Exception as e:
            print(f"保存工作簿失败：{e}")
            return False

        print(f"insert_columns 函数执行完毕")
        print("-" * 30)
        return True

    except Exception as e:
        print(f"insert_columns函数执行过程中发生未知错误：{e}")
        return False
    finally:
        print(f"insert_columns 函数执行完毕")
        print("-" * 30)

'''