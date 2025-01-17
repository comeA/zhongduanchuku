import openpyxl

def copy_sheet_data(source_filepath, source_sheet, dest_filepath, result_sheet):
    """复制工作表数据"""
    print("-" * 30)
    print("开始执行 copy_sheet_data 函数")
    print(f"源文件路径：{source_filepath}, 源工作表：{source_sheet}")
    print(f"目标文件路径：{dest_filepath}, 结果工作表：{result_sheet}")
    try:
        try:
            source_wb = openpyxl.load_workbook(source_filepath)
            source_sh = source_wb[source_sheet]
        except FileNotFoundError:
            print(f"错误：文件 {source_filepath} 未找到。")
            return False, None
        except KeyError:
            print(f"错误：工作表 {source_sheet} 未找到。")
            return False, None
        except Exception as e:
            print(f"读取源文件失败：{e}")
            return False, None

        try:
            dest_wb = openpyxl.load_workbook(dest_filepath) # 尝试加载目标文件
        except FileNotFoundError:
            dest_wb = openpyxl.Workbook() # 如果文件不存在，则创建新工作簿
        except Exception as e:
            print(f"读取/创建目标文件失败：{e}")
            return False, None

        if result_sheet in dest_wb.sheetnames:
            dest_wb.remove(dest_wb[result_sheet]) #删除已存在的sheet
            print(f"目标文件已存在sheet:{result_sheet},已删除")
        dest_sh = dest_wb.create_sheet(result_sheet)

        for row in source_sh.iter_rows():
            dest_sh.append([cell.value for cell in row])

        try:
            dest_wb.save(dest_filepath)
        except Exception as e:
            print(f"保存目标文件失败：{e}")
            return False, None

        print("工作表数据复制成功！")
        return True, result_sheet
    except Exception as e:
        print(f"复制工作表数据时发生未知错误：{e}")
        return False, None
    finally:
        print("copy_sheet_data 函数执行完毕")
        print("-" * 30)