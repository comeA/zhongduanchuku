import openpyxl

def copy_data_to_excel(df_column, filename, sheetname, header_name):
    """将 DataFrame 的一列数据复制到 Excel 文件的指定工作表"""
    try:
        values = df_column.dropna().tolist()
        wb = openpyxl.load_workbook(filename)
        try:
            ws = wb[sheetname]  # 使用用户指定的工作表名称
        except KeyError:
            print(f"{filename} 文件中不存在名为 {sheetname} 的工作表，请检查！")
            return False
        # 清空数据，保留表头
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        ws.cell(row=1, column=1).value = header_name #写入表头
        for i, value in enumerate(values):
            ws.cell(row=i + 2, column=1).value = value
        wb.save(filename)
        print(f"数据已成功复制到 {filename} 的 {sheetname} 工作表")
        return True
    except FileNotFoundError as e:
        print(f"错误：{filename} 文件未找到：{e}")
        return False
    except Exception as e:
        print(f"复制数据到 {filename} 文件时发生错误：{e}")
        return False